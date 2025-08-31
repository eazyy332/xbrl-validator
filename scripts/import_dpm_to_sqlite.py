import argparse
import csv
import os
import re
import shutil
import subprocess
import sqlite3
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Iterable, List, Tuple

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


def sanitize_table_name(name: str) -> str:
    base = re.sub(r"[^A-Za-z0-9_]+", "_", name)
    base = re.sub(r"_+", "_", base).strip("_")
    if not base:
        base = "table"
    # SQLite: avoid starting with digits
    if base[0].isdigit():
        base = f"t_{base}"
    return base.lower()


def read_csv_records(path: Path) -> Tuple[List[str], Iterable[List[str]]]:
    # Keep file open for the duration of iteration; close when done
    f = path.open("r", encoding="utf-8-sig", newline="")
    reader = csv.reader(f)
    headers = next(reader)

    def row_iter():
        try:
            for row in reader:
                yield row
        finally:
            try:
                f.close()
            except Exception:
                pass

    return headers, row_iter()


def read_xlsx_records(path: Path) -> Iterable[Tuple[str, List[str], Iterable[List[str]]]]:
    if load_workbook is None:
        return []
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    for ws in wb.worksheets:
        # Assume first row is header
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            continue
        def row_iter():
            for row in rows_iter:
                yield ["" if v is None else str(v) for v in row]
        yield (ws.title, headers, row_iter())


def ensure_table(conn: sqlite3.Connection, table: str, headers: List[str]) -> None:
    cols = [sanitize_table_name(h or f"col{i}") for i, h in enumerate(headers)]
    cols_unique = []
    seen = set()
    for c in cols:
        nn = c
        k = 1
        while nn in seen:
            nn = f"{c}_{k}"
            k += 1
        cols_unique.append(nn)
        seen.add(nn)
    cols_sql = ", ".join(f'"{c}" TEXT' for c in cols_unique)
    conn.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({cols_sql});')


def bulk_insert(conn: sqlite3.Connection, table: str, headers: List[str], rows: Iterable[List[str]], batch: int = 1000) -> int:
    cols = [sanitize_table_name(h or f"col{i}") for i, h in enumerate(headers)]
    placeholders = ",".join(["?"] * len(cols))
    cols_sql = ",".join([f'"{c}"' for c in cols])
    sql = f'INSERT INTO "{table}" ({cols_sql}) VALUES ({placeholders})'
    count = 0
    buf: List[List[str]] = []
    for r in rows:
        # Pad/truncate row to headers length
        if len(r) < len(cols):
            r = r + ([None] * (len(cols) - len(r)))
        elif len(r) > len(cols):
            r = r[: len(cols)]
        buf.append(r)
        if len(buf) >= batch:
            conn.executemany(sql, buf)
            conn.commit()
            count += len(buf)
            buf.clear()
    if buf:
        conn.executemany(sql, buf)
        conn.commit()
        count += len(buf)
    return count


def export_accdb_to_csvs(accdb_path: Path, out_dir: Path) -> None:
    mdb_tables = shutil.which("mdb-tables")
    mdb_export = shutil.which("mdb-export")
    if not mdb_tables or not mdb_export:
        print("[warn] mdbtools not found (mdb-tables/mdb-export). Install via: brew install mdbtools")
        return
    # List tables (one per line)
    res = subprocess.run([mdb_tables, "-1", str(accdb_path)], capture_output=True, text=True)
    if res.returncode != 0:
        print("[warn] mdb-tables failed:", res.stderr.strip())
        return
    table_names = [t.strip() for t in res.stdout.splitlines() if t.strip()]
    for tname in table_names:
        # Export CSV for each table
        csv_path = out_dir / f"{sanitize_table_name(tname)}.csv"
        with csv_path.open("w", encoding="utf-8") as f:
            exp = subprocess.run(
                [mdb_export, str(accdb_path), tname],
                stdout=f,
                text=True,
            )
        if exp.returncode == 0:
            print(f"[accdb] exported {tname} -> {csv_path.name}")
        else:
            print(f"[accdb] export failed for {tname}")


def import_zip_to_sqlite(zip_path: Path, sqlite_path: Path, schema_prefix: str) -> None:
    tmpdir = Path(tempfile.mkdtemp(prefix="dpm_zip_"))
    try:
        with zipfile.ZipFile(str(zip_path), "r") as zf:
            zf.extractall(tmpdir)
        conn = sqlite3.connect(str(sqlite_path))
        try:
            # Scan for CSV and XLSX
            for p in tmpdir.rglob("*"):
                if p.is_file() and p.suffix.lower() == ".csv":
                    table = sanitize_table_name(f"{schema_prefix}_{p.stem}")
                    headers, rows = read_csv_records(p)
                    ensure_table(conn, table, headers)
                    n = bulk_insert(conn, table, headers, rows)
                    print(f"[csv] {p.name} -> {table}: {n} rows")
                elif p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm"):
                    if load_workbook is None:
                        print(f"[xlsx] skipped (openpyxl not available): {p}")
                        continue
                    for sheet_name, headers, rows in read_xlsx_records(p):
                        table = sanitize_table_name(f"{schema_prefix}_{p.stem}_{sheet_name}")
                        ensure_table(conn, table, headers)
                        n = bulk_insert(conn, table, headers, rows)
                        print(f"[xlsx] {p.name}!{sheet_name} -> {table}: {n} rows")
                elif p.is_file() and p.suffix.lower() == ".accdb":
                    # Attempt export via mdbtools then re-scan for CSVs
                    export_dir = tmpdir / "_accdb_csv"
                    export_dir.mkdir(exist_ok=True)
                    export_accdb_to_csvs(p, export_dir)
                    for c in export_dir.glob("*.csv"):
                        table = sanitize_table_name(f"{schema_prefix}_{c.stem}")
                        headers, rows = read_csv_records(c)
                        ensure_table(conn, table, headers)
                        n = bulk_insert(conn, table, headers, rows)
                        print(f"[accdb-csv] {c.name} -> {table}: {n} rows")
        finally:
            conn.close()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def main() -> int:
    ap = argparse.ArgumentParser(description="Import EBA DPM (CSV/XLSX) into SQLite")
    ap.add_argument("--zip", required=True, help="Path to a DPM zip containing CSV/XLSX")
    ap.add_argument("--sqlite", required=True, help="Output SQLite file path (will be created/updated)")
    ap.add_argument("--schema", required=True, help="Schema prefix (e.g., dpm35_20)")
    args = ap.parse_args()

    zip_path = Path(args.zip)
    sqlite_path = Path(args.sqlite)
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)

    if not zip_path.exists():
        print(f"[error] Zip not found: {zip_path}")
        return 2

    import_zip_to_sqlite(zip_path, sqlite_path, args.schema)
    print(f"[done] SQLite updated: {sqlite_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


