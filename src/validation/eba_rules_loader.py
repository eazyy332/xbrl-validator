from __future__ import annotations

from pathlib import Path
from typing import Dict, Any, List, Optional
import json
import hashlib

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def load_rules_from_excel(xlsx_path: str, cache_json: Optional[str] = None) -> Dict[str, Any]:
    """Load EBA Filing Rules from Excel and return a normalized dict.

    The Excel is large (~12MB, 136k rows). We stream rows and keep minimal fields:
    - rule_id, framework, table, condition, severity, message_template, applicability hints
    If cache_json provided, we also write a compact JSON cache for faster reuse.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed; cannot read EBA rules Excel")

    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        raise FileNotFoundError(f"EBA rules Excel not found: {xlsx}")

    wb = openpyxl.load_workbook(filename=str(xlsx), read_only=True, data_only=True)
    # Heuristic: find first sheet containing 'Rule' and many rows
    ws = None
    for name in wb.sheetnames:
        if "rule" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        # fallback to active
        ws = wb.active

    # Map headers
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = [(_norm(c).lower()) for c in row]
        break
    if not header_row:
        raise RuntimeError("EBA rules Excel: header row not found")

    def col(name_candidates: List[str]) -> int:
        for cand in name_candidates:
            if cand in header_row:
                return header_row.index(cand)
        return -1

    idx_rule = col(["rule_id", "rule id", "id"])
    idx_sev = col(["severity", "level"])
    idx_msg = col(["message", "message_template", "text"])
    idx_cond = col(["condition", "expression", "formula", "rule_expression"])
    idx_table = col(["table", "template", "table_id"])
    idx_framework = col(["framework", "version", "eba_version"])
    idx_app = col(["applicability", "filing indicator", "filing_indicator", "applicability notes"]) 
    idx_code = col(["code", "rule_code", "eba_code"]) 
    idx_prereq = col(["prerequisite", "precondition", "prereq"]) 
    # Lifecycle columns (various historical names)
    idx_active = col(["active", "status", "is_active", "enabled"]) 
    idx_valid_from = col(["valid from", "valid_from", "start", "start date", "from"]) 
    idx_valid_to = col(["valid to", "valid_to", "end", "end date", "to"]) 

    rules: List[Dict[str, Any]] = []
    row_number = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            row_number += 1
            rid = _norm(row[idx_rule]) if idx_rule >= 0 else ""
            if not rid:
                continue
            # Normalize lifecycle
            status_val = _norm(row[idx_active]) if idx_active >= 0 else ""
            is_active = True
            if status_val:
                sv = status_val.lower()
                # Treat typical values
                is_active = sv in ("1", "true", "yes", "active", "enabled", "y")
                if sv in ("inactive", "disabled", "no", "false", "0"):
                    is_active = False
            vfrom = _norm(row[idx_valid_from]) if idx_valid_from >= 0 else ""
            vto = _norm(row[idx_valid_to]) if idx_valid_to >= 0 else ""

            rules.append({
                "id": rid,
                "severity": (_norm(row[idx_sev]) if idx_sev >= 0 else "").upper() or "ERROR",
                "message": _norm(row[idx_msg]) if idx_msg >= 0 else "",
                "condition": _norm(row[idx_cond]) if idx_cond >= 0 else "",
                "table": _norm(row[idx_table]) if idx_table >= 0 else "",
                "framework": _norm(row[idx_framework]) if idx_framework >= 0 else "",
                "applicability": _norm(row[idx_app]) if idx_app >= 0 else "",
                "code": _norm(row[idx_code]) if idx_code >= 0 else "",
                "prereq": _norm(row[idx_prereq]) if idx_prereq >= 0 else "",
                # Lifecycle
                "active": bool(is_active),
                "valid_from": vfrom,
                "valid_to": vto,
                # Provenance
                "_sheet": str(getattr(ws, 'title', '')),
                "_row": int(row_number),
            })
        except Exception:
            continue

    data = {"count": len(rules), "rules": rules}

    if cache_json:
        p = Path(cache_json)
        p.parent.mkdir(parents=True, exist_ok=True)
        with p.open("w", encoding="utf-8") as f:
            json.dump(data, f)

    return data


def load_cached_rules(cache_json: str) -> Optional[Dict[str, Any]]:
    p = Path(cache_json)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def excel_sha256(xlsx_path: str) -> str:
    """Return a stable sha256 for one or more Excel paths.

    Supports semicolon or pipe separated lists to fingerprint combinations
    like "base.xlsx;overlay1.xlsx;overlay2.xlsx" deterministically.
    """
    # Split by common separators and normalize
    parts = [s.strip() for s in str(xlsx_path).split(";") if s.strip()]
    # Also support '|' as separator
    if len(parts) == 1 and ("|" in str(xlsx_path)):
        parts = [s.strip() for s in str(xlsx_path).split("|") if s.strip()]
    # Remove duplicates and sort for stability
    paths: List[Path] = sorted({Path(p) for p in parts}, key=lambda p: str(p)) if parts else [Path(xlsx_path)]
    h = hashlib.sha256()
    for p in paths:
        if not p.exists():
            # Include missing path text to still get a unique fingerprint
            h.update(str(p).encode("utf-8"))
            continue
        with p.open('rb') as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
    return h.hexdigest()


def build_rules_index(data: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """Build an index of rules keyed by normalized lower-case table id.
    The special key '*' contains global rules that have no table.
    Each rule contains a translated condition string under 'cond_expr'.
    """
    from .eba_rules import _normalize_table_id, _translate_condition_to_expr  # local import to avoid cycle at module import

    idx: Dict[str, List[Dict[str, Any]]] = {}
    for r in data.get("rules", []) or []:
        try:
            rtab = _normalize_table_id((r.get("table") or "").strip())
            key = (rtab.lower() if rtab else "*")
            item = dict(r)
            try:
                item["cond_expr"] = _translate_condition_to_expr((item.get("condition") or "").strip())
            except Exception:
                item["cond_expr"] = (item.get("condition") or "").strip()
            idx.setdefault(key, []).append(item)
        except Exception:
            continue
    return idx


def save_rules_index(index: Dict[str, List[Dict[str, Any]]], out_json: str) -> None:
    p = Path(out_json)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open('w', encoding='utf-8') as f:
        json.dump(index, f)


def load_rules_index(path: str) -> Optional[Dict[str, List[Dict[str, Any]]]]:
    p = Path(path)
    if not p.exists():
        return None
    try:
        data = json.loads(p.read_text(encoding='utf-8'))
        # typing hint
        return data  # type: ignore[return-value]
    except Exception:
        return None


